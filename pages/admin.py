'''
    ADMIN PAGE FOR NET STATISTICS 
'''


## MODULE IMPORTS ##
import streamlit as st
import pandas as pd
import altair as alt
import io
import matplotlib.pyplot as plt
import openpyxl as oxl

from google.cloud import firestore
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from tempfile import NamedTemporaryFile
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime


## AUTHORIZATION CHECK ##
# try:
#     if st.session_state['logged_in_admin'] == False:
#         st.switch_page('app.py')
# except:
#     st.switch_page('app.py')


## STREAMLIT PAGE STYLESHEET ##
st.set_page_config(layout='wide', initial_sidebar_state='collapsed')

st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        display: none;
    }
    </style>
    """, unsafe_allow_html=True)

st.markdown(
    """
    <style>
        [data-testid="collapsedControl"]{
            display: none
        }S
    </style>
    """, unsafe_allow_html=True)


## GLOBAL FIELD ##
db = firestore.Client.from_service_account_json(".streamlit/firebase_key.json")


label_counts = {}
docs = db.collection("trains").get()

new_class_names = {
    'assault': '폭행',
    'fainting': '실신',
    'property_damage': '기물파손',
    'theft': '절도',
    'merchant': '잡상인',
    'spy_camera': '몰래카메라'
}

for doc in docs:
    data = doc.to_dict()
    data_kor = {new_class_names.get(key, key): value for key, value in data.items()}
    for key, value in data_kor.items():
        if key in label_counts:
            label_counts[key] += value
        else:
            label_counts[key] = value

class_name = list(label_counts.keys())
total_value = list(label_counts.values())


def render_chart(label_counts):
    df = pd.DataFrame({'객체': class_name, '총 탐지수': label_counts})
    
    chart = alt.Chart(df).mark_bar().encode(
        x=alt.X('객체:N', axis=alt.Axis(labelAngle=0, title='객체')),
        y=alt.Y('총 탐지수:Q', axis=alt.Axis(title='총 탐지수'), scale=alt.Scale(domain=[0, max(label_counts) + 1]))
    ).properties(
        width=alt.Step(40)
    )
    
    s1 = dict(selector='th', props=[('text-align', 'center')])
    s2 = dict(selector='td', props=[('text-align', 'center')])
    table = df.style.set_table_styles([s1,s2]).hide(axis=0).to_html()

    col1, col2, col3 = st.columns([2, 0.3, 10])

    with col1:
        st.write(f'{table}', unsafe_allow_html=True)

    with col2:
        pass

    with col3:
        st.altair_chart(chart, use_container_width=True)

st.header('CCTV 이상행동 탐지 통계', divider='rainbow')


with st.container(border=True):
    render_chart(total_value)

current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 엑셀 파일 생성
output = io.BytesIO()
with pd.ExcelWriter(output, engine='openpyxl') as writer:
    x = 0
    for i in range(0, len(docs)):
        data = docs[i].to_dict()
        data_kor = {new_class_names.get(key, key): value for key, value in data.items()}
        df = pd.DataFrame({'객체': data_kor.keys(), '탐지수': data_kor.values()} )
        total_df = pd.DataFrame({'객체': class_name, '탐지수': total_value})
        if i == 0:
            x = 10
        else:
            x += 18
            
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow = x, startcol=1)

        if i + 2 == len(docs) + 1:
            total_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow = 18*len(docs)+8, startcol=1)
        
        # 차트 생성 및 설정
        chart1 = BarChart()
        chart1.style = 10
        chart1.title = '탐지수'
        chart1.x_axis.delete = False
        chart1.y_axis.delete = False
        chart1.y_axis.majorUnit = 1

        # 엑셀 워크북과 워크시트 가져오기
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        worksheet.column_dimensions['B'].width = 10
        worksheet.merge_cells('E2:I4')
        #
        cell = worksheet['E2']
        cell.value = f'{current_time.split("-")[1]}월 탐지 결과 보고서'
        cell.font = Font(name = 'HY견명조', size = 18, italic = True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('K5:M5')
        cell_k5_m5 = worksheet['K5']
        cell_k5_m5.value = current_time
        cell_k5_m5.alignment = Alignment(horizontal='center', vertical='center')

        
        from openpyxl.styles import Alignment

        # 데이터와 스타일 설정
        # 차량 1
        worksheet.merge_cells('B10:C10')
        cell_b10 = worksheet['B10']
        cell_b10.value = '차량 1'
        cell_b10.alignment = Alignment(horizontal='center', vertical='center')

        # 차량 2
        worksheet.merge_cells('B28:C28')
        cell_b25 = worksheet['B28']
        cell_b25.value = '차량 2'
        cell_b25.alignment = Alignment(horizontal='center', vertical='center')

        # 차량 3
        worksheet.merge_cells('B46:C46')
        cell_b40 = worksheet['B46']
        cell_b40.value = '차량 3'
        cell_b40.alignment = Alignment(horizontal='center', vertical='center')

        # 누적
        worksheet.merge_cells('B62:C62')
        cell_a52_b52 = worksheet['B62']
        cell_a52_b52.value = '누적'
        cell_a52_b52.alignment = Alignment(horizontal='center', vertical='center')


        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        #일자
        for row in worksheet.iter_rows(min_row=5, max_row=5, min_col=11, max_col=13):
            for cell in row:
                cell.border = thin_border

        #타이틀
        for row in worksheet.iter_rows(min_row=2, max_row=4, min_col=5, max_col=9):
            for cell in row:
                cell.border = thin_border

        border_ranges = [
            (10, 17),
            (28, 35),
            (46, 53),
            (62, 69)
        ]

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for start_row, end_row in border_ranges:
            for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=2, max_col=3):
                for cell in row:
                    cell.border = thin_border


        
        
        ## car_1 일떄 x = 10
        ## car_2 일떄 x = 18
        ## car_3 일떄 x = 36
        ## len(df) = 6

        data_range = len(df) + x + 1
        data = Reference(worksheet, min_col=3, min_row=x+1, max_row=data_range, max_col=3)
        categories = Reference(worksheet, min_col=2, min_row=x+2, max_row=data_range, max_col=2)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)

        chart2 = BarChart()
        chart2.style = 10
        chart2.title = '전체 탐지수'
        chart2.x_axis.delete = False
        chart2.y_axis.delete = False
        chart2.y_axis.majorUnit = 1

        t_data=Reference(worksheet, min_col=3, min_row=18*len(docs)+9, max_row=18*len(docs)+15, max_col=3)
        t_categories = Reference(worksheet,min_col=2, min_row=18*len(docs)+10, max_row=18*len(docs)+15, max_col=2)
        chart2.add_data(t_data,titles_from_data=True)
        chart2.set_categories(t_categories)


        # 엑셀 워크시트에 차트 삽입            
        worksheet.add_chart(chart1, f'F{x}')
        worksheet.add_chart(chart2,f'F{18*len(docs)+8}')

            
# 파일 포인터를 처음으로 되돌리고 다운로드 버튼 생성
output.seek(0)
st.download_button(label="저장", data=output, file_name=f'{current_time.split("-")[1]}월 이상행동 탐지 보고서.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# 가장 마지막으로 할 스타일 : 각 데이터 표 위에 이름 달아주기, 그래프 포함